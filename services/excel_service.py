# -*- coding: utf-8 -*-
"""Excel okuma/yazma: sabit grid (gün sütun, saat satır)."""

from pathlib import Path
from typing import Any, Dict, Optional

from openpyxl import load_workbook
from openpyxl.styles import Alignment

from constants import COL_TO_DAY, DAY_TO_COL, GUNLER, ROW_TO_TIME, SAYFA_ADI, TIME_TO_ROW


def _parse_entry(entry: Any) -> tuple[str, bool]:
    if isinstance(entry, dict):
        return entry.get("text", ""), bool(entry.get("done", False))
    if entry is None:
        return "", False
    return str(entry), False


class ExcelService:
    """Koçluk çizelgesi Excel dosyası: okuma (grid doldurma), yazma (kaydetme)."""

    def __init__(self, sheet_name: Optional[str] = None):
        self.sheet_name = sheet_name or SAYFA_ADI

    def load_program_from_file(self, filepath: str) -> Dict[str, Dict[str, Any]]:
        """
        Excel dosyasından programı okur. Sayfa adı SAYFA_ADI.
        Döner: program[gun][saat] = {"text": str, "done": bool}
        """
        program = {g: {} for g in GUNLER}
        wb = load_workbook(filepath)
        if self.sheet_name not in wb.sheetnames:
            wb.close()
            return program
        ws = wb[self.sheet_name]
        min_row = min(ROW_TO_TIME.keys()) if ROW_TO_TIME else 3
        max_row = max(ROW_TO_TIME.keys()) if ROW_TO_TIME else 16
        for row in range(min_row, max_row + 1):
            for col in range(2, 9):
                deger = ws.cell(row=row, column=col).value
                if deger is None:
                    continue
                gun = COL_TO_DAY.get(col)
                saat = ROW_TO_TIME.get(row)
                if gun and saat:
                    program[gun][saat] = {"text": str(deger), "done": False}
        wb.close()
        return program

    def save_program_to_file(self, filepath: str, program: Dict[str, Dict[str, Any]]) -> None:
        """
        Programı Excel dosyasına yazar. Mevcut sayfayı temizleyip programı yazar.
        Dosya açıksa PermissionError fırlatır.
        """
        wb = load_workbook(filepath)
        if self.sheet_name not in wb.sheetnames:
            wb.close()
            raise ValueError(f"'{self.sheet_name}' sayfası bulunamadı.")
        ws = wb[self.sheet_name]
        min_row = min(ROW_TO_TIME.keys()) if ROW_TO_TIME else 3
        max_row = max(ROW_TO_TIME.keys()) if ROW_TO_TIME else 16
        for row in range(min_row, max_row + 1):
            for col in range(2, 9):
                cell = ws.cell(row=row, column=col)
                cell.value = None
                cell.alignment = Alignment(wrap_text=True)
        for gun, saat_dict in program.items():
            col = DAY_TO_COL.get(gun)
            if col is None:
                continue
            for saat, entry in saat_dict.items():
                row = TIME_TO_ROW.get(saat)
                if row is None:
                    continue
                text, _ = _parse_entry(entry)
                cell = ws.cell(row=row, column=col)
                cell.value = text
                cell.alignment = Alignment(wrap_text=True)
        for row in range(min_row, max_row + 1):
            dolu = any(ws.cell(row=row, column=col).value for col in range(2, 9))
            ws.row_dimensions[row].height = 35 if dolu else 18
        wb.save(filepath)
        wb.close()
